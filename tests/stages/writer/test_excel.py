"""ExcelWriter unit tests against a synthetic WitrineSnapshot."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from edx.stages.writer.excel import (
    DECIMAL_NUMBER_FORMAT,
    INTEGER_NUMBER_FORMAT,
    EventExportRow,
    ExcelWriter,
    MetaSnapshot,
    MetricExportRow,
    QAIssueExportRow,
    TickerExportRow,
    WitrineSnapshot,
)


def _snapshot() -> WitrineSnapshot:
    return WitrineSnapshot(
        metrics=[
            MetricExportRow(
                ticker="SBER",
                reporting_date="2025-12-31",
                period_type="FY",
                reporting_standard="IFRS",
                metric_name="revenue",
                value=1_500_000.0,
                currency="RUB",
                unit="ones",
                qa_warning=None,
                source_publication_url="https://example.test/sber-fy25",
            ),
            MetricExportRow(
                ticker="SBER",
                reporting_date="2025-12-31",
                period_type="FY",
                reporting_standard="IFRS",
                metric_name="net_income",
                value=42_000.5,
                currency="RUB",
                unit="ones",
                qa_warning=None,
                source_publication_url="https://example.test/sber-fy25",
            ),
        ],
        events=[
            EventExportRow(
                ticker="SBER",
                event_date="2026-04-28",
                publication_date="2026-04-28",
                event_type="dividends",
                summary="Дивиденды 22,5 руб.",
                key_params_json='{"per_share_rub": 22.5}',
                source_url="https://example.test/sber-divs",
            )
        ],
        qa_issues=[
            QAIssueExportRow(
                ticker="SBER",
                publication_id="pub-1",
                code="incomplete",
                message="Извлечено 2 из 5 показателей.",
                created_at="2026-05-01T00:00:00+00:00",
            )
        ],
        tickers=[
            TickerExportRow(
                ticker="SBER",
                name="Sberbank",
                profile="bank",
                e_disclosure_id="3043",
            ),
            TickerExportRow(
                ticker="LKOH",
                name="Lukoil",
                profile="non_bank",
                e_disclosure_id="17",
                use_vision_extraction=True,  # Patch 34: surfaced in Excel
            ),
        ],
        meta=MetaSnapshot(
            last_updated_at="2026-05-01T00:00:00+00:00",
            pipeline_version="0.1.0",
            tickers_count=3,
            metrics_rows=2,
            events_rows=1,
            incomplete_publications=1,
            failed_publications=0,
        ),
    )


def test_excel_writer_round_trips_all_sheets(tmp_path: Path) -> None:
    target = tmp_path / "mart" / "e-disclosure.xlsx"
    ExcelWriter().write(target, _snapshot())
    assert target.is_file()
    wb = load_workbook(target)
    # Patch 19: dedicated ``tickers`` sheet with the issuer profile column.
    assert set(wb.sheetnames) == {
        "metrics",
        "events",
        "tickers",
        "meta",
        "qa_issues",
    }
    tickers = wb["tickers"]
    headers = [c.value for c in tickers[1]]
    # Patch 34: tickers sheet now carries the use_vision_extraction column.
    assert headers == [
        "ticker",
        "name",
        "profile",
        "e_disclosure_id",
        "use_vision_extraction",
    ]
    profiles = {
        tickers.cell(row=r, column=1).value: tickers.cell(row=r, column=3).value
        for r in range(2, tickers.max_row + 1)
    }
    assert profiles == {"SBER": "bank", "LKOH": "non_bank"}
    # Patch 34: column 5 = use_vision_extraction. SBER default False, LKOH True.
    vision_flags = {
        tickers.cell(row=r, column=1).value: tickers.cell(row=r, column=5).value
        for r in range(2, tickers.max_row + 1)
    }
    assert vision_flags == {"SBER": False, "LKOH": True}


def test_metrics_sheet_headers_and_values(tmp_path: Path) -> None:
    target = tmp_path / "e-disclosure.xlsx"
    ExcelWriter().write(target, _snapshot())
    wb = load_workbook(target)
    metrics = wb["metrics"]
    headers = [c.value for c in metrics[1]]
    assert headers == [
        "ticker",
        "reporting_date",
        "period_type",
        "reporting_standard",
        "metric_name",
        "value",
        "currency",
        "unit",
        "qa_warning",
        "source_publication_url",
    ]
    assert metrics["A2"].value == "SBER"
    assert metrics["E2"].value == "revenue"
    assert metrics["F2"].value == 1_500_000.0
    # Bold headers.
    assert metrics["A1"].font.bold is True
    # Freeze pane at A2.
    assert metrics.freeze_panes == "A2"


def test_metrics_number_format_branches_on_integer_vs_float(
    tmp_path: Path,
) -> None:
    target = tmp_path / "e-disclosure.xlsx"
    ExcelWriter().write(target, _snapshot())
    wb = load_workbook(target)
    metrics = wb["metrics"]
    integer_cell = metrics["F2"]  # 1_500_000.0 — integer
    decimal_cell = metrics["F3"]  # 42_000.5 — non-integer
    assert integer_cell.number_format == INTEGER_NUMBER_FORMAT
    assert decimal_cell.number_format == DECIMAL_NUMBER_FORMAT


def test_events_sheet_present(tmp_path: Path) -> None:
    target = tmp_path / "e-disclosure.xlsx"
    ExcelWriter().write(target, _snapshot())
    wb = load_workbook(target)
    events = wb["events"]
    headers = [c.value for c in events[1]]
    assert headers == [
        "ticker",
        "event_date",
        "publication_date",
        "event_type",
        "summary",
        "key_params_json",
        "source_url",
    ]
    assert events["D2"].value == "dividends"


def test_meta_sheet_key_value_pairs(tmp_path: Path) -> None:
    target = tmp_path / "e-disclosure.xlsx"
    ExcelWriter().write(target, _snapshot())
    wb = load_workbook(target)
    meta = wb["meta"]
    pairs = {str(meta.cell(row=r, column=1).value): meta.cell(row=r, column=2).value
             for r in range(2, meta.max_row + 1)}
    assert pairs["pipeline_version"] == "0.1.0"
    assert pairs["tickers_count"] == 3
    assert pairs["metrics_rows"] == 2
    assert pairs["events_rows"] == 1
    assert pairs["incomplete_publications"] == 1
    assert pairs["failed_publications"] == 0


def test_qa_issues_sheet_present(tmp_path: Path) -> None:
    target = tmp_path / "e-disclosure.xlsx"
    ExcelWriter().write(target, _snapshot())
    wb = load_workbook(target)
    sheet = wb["qa_issues"]
    headers = [c.value for c in sheet[1]]
    assert headers == [
        "ticker",
        "publication_id",
        "code",
        "message",
        "created_at",
    ]
    assert sheet["C2"].value == "incomplete"


def test_atomic_write_keeps_prior_file_on_failure(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "e-disclosure.xlsx"
    # First, write a valid file with the original snapshot.
    ExcelWriter().write(target, _snapshot())
    original_bytes = target.read_bytes()

    # Force failure inside Workbook.save before os.replace runs.
    import edx.stages.writer.excel as writer_module

    def _failing_save(self: object, target_path: object) -> None:
        # Touch a partial .tmp to mimic openpyxl writing then crashing.
        Path(str(target_path)).write_bytes(b"partial")
        raise RuntimeError("simulated failure")

    monkeypatch.setattr(writer_module.Workbook, "save", _failing_save)

    with pytest.raises(RuntimeError, match="simulated"):
        ExcelWriter().write(target, _snapshot())

    # Original file unchanged; partial .tmp cleaned up.
    assert target.read_bytes() == original_bytes
    assert not target.with_name(target.name + ".tmp").exists()


def test_empty_snapshot_writes_only_headers(tmp_path: Path) -> None:
    target = tmp_path / "empty.xlsx"
    ExcelWriter().write(
        target,
        WitrineSnapshot(
            metrics=[],
            events=[],
            qa_issues=[],
            meta=None,
        ),
    )
    wb = load_workbook(target)
    metrics = wb["metrics"]
    # Only the header row.
    assert metrics.max_row == 1
    assert metrics["A1"].value == "ticker"
