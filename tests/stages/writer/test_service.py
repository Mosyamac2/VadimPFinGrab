"""WriterService end-to-end on a tmp SQLite + tmp output path."""

from __future__ import annotations

from contextlib import closing
from pathlib import Path

import pytest
from openpyxl import load_workbook

from edx.config import TickerEntry, TickersConfig
from edx.stages.writer.service import WriterService
from edx.storage import (
    Database,
    DocumentInput,
    DocumentsRepo,
    EventInput,
    EventsRepo,
    MetricInput,
    MetricsRepo,
    PublicationsRepo,
    QAIssuesRepo,
    TickersRepo,
)


def _seed(db: Database) -> None:
    with closing(db.connect()) as conn:
        TickersRepo(db, conn).upsert_from_config(
            [
                TickerEntry(ticker="SBER", e_disclosure_id="1", name="Sberbank"),
                TickerEntry(ticker="GAZP", e_disclosure_id="2", name="Gazprom"),
            ]
        )
        pubs = PublicationsRepo(db, conn)
        docs = DocumentsRepo(db, conn)
        metrics_repo = MetricsRepo(db, conn)
        events_repo = EventsRepo(db, conn)

        # Two report publications, 5 metrics each.
        for ticker, pub_id in [("SBER", "rep-sber"), ("GAZP", "rep-gazp")]:
            pubs.upsert_discovered(
                publication_id=pub_id,
                ticker=ticker,
                publication_type="report",
                publication_date="2026-04-01",
                source_url=f"https://example.test/{pub_id}",
            )
            for status in (
                "downloaded",
                "unpacked",
                "classified",
                "extracted",
                "validated",
            ):
                pubs.mark_status(pub_id, status)  # type: ignore[arg-type]
            docs.add_documents(
                pub_id,
                [DocumentInput(relative_path="r.pdf", file_hash=f"h-{pub_id}")],
            )
            doc_id = docs.list_for_publication(pub_id)[0].document_id
            rows = [
                MetricInput(
                    ticker=ticker,
                    reporting_date="2025-12-31",
                    period_type="FY",
                    reporting_standard="IFRS",
                    metric_name=name,
                    value=value,
                    currency="RUB",
                    unit="ones",
                    source_document_id=doc_id,
                )
                for name, value in [
                    ("revenue", 1_000_000.0),
                    ("ebitda", 300_000.0),
                    ("net_income", 100_000.0),
                    ("total_assets", 5_000_000.0),
                    ("total_debt", 2_000_000.0),
                ]
            ]
            metrics_repo.replace_for_publication(pub_id, rows)

        # One event.
        pubs.upsert_discovered(
            publication_id="ev-1",
            ticker="SBER",
            publication_type="event",
            publication_date="2026-04-28",
            source_url="https://example.test/ev-1",
        )
        for status in (
            "downloaded",
            "unpacked",
            "classified",
            "extracted",
            "validated",
        ):
            pubs.mark_status("ev-1", status)  # type: ignore[arg-type]
        events_repo.upsert_event(
            EventInput(
                ticker="SBER",
                event_date="2026-04-28",
                publication_date="2026-04-28",
                event_type="dividends",
                summary="Дивиденды 22,5 руб.",
                key_params_json='{"per_share_rub": 22.5}',
                source_url="https://example.test/ev-1",
                source_publication_id="ev-1",
            )
        )


_TICKERS_CONFIG = TickersConfig(
    tickers=[
        TickerEntry(
            ticker="SBER",
            e_disclosure_id="1",
            name="Sberbank",
            profile="bank",
        ),
        TickerEntry(
            ticker="GAZP",
            e_disclosure_id="2",
            name="Gazprom",
            profile="non_bank",
        ),
    ]
)


@pytest.fixture
def workspace(tmp_path: Path) -> tuple[Database, Path]:
    db = Database(tmp_path / "state.sqlite")
    db.migrate()
    _seed(db)
    excel_path = tmp_path / "output" / "e-disclosure.xlsx"
    return db, excel_path


def test_writer_emits_expected_sheet_counts(
    workspace: tuple[Database, Path],
) -> None:
    db, excel_path = workspace
    with closing(db.connect()) as conn:
        service = WriterService(
            PublicationsRepo(db, conn),
            MetricsRepo(db, conn),
            EventsRepo(db, conn),
            QAIssuesRepo(db, conn),
            TickersRepo(db, conn),
            tickers_config=_TICKERS_CONFIG,
            excel_path=excel_path,
        )
        out = service.run()

    assert out == excel_path
    wb = load_workbook(out)
    metrics = wb["metrics"]
    events = wb["events"]
    meta = wb["meta"]
    # Header + 10 data rows for metrics; header + 1 for events.
    assert metrics.max_row == 11
    assert events.max_row == 2

    pairs = {
        str(meta.cell(row=r, column=1).value): meta.cell(row=r, column=2).value
        for r in range(2, meta.max_row + 1)
    }
    assert pairs["tickers_count"] == 2
    assert pairs["metrics_rows"] == 10
    assert pairs["events_rows"] == 1


def test_writer_marks_validated_publications_as_written(
    workspace: tuple[Database, Path],
) -> None:
    db, excel_path = workspace
    with closing(db.connect()) as conn:
        service = WriterService(
            PublicationsRepo(db, conn),
            MetricsRepo(db, conn),
            EventsRepo(db, conn),
            QAIssuesRepo(db, conn),
            TickersRepo(db, conn),
            tickers_config=_TICKERS_CONFIG,
            excel_path=excel_path,
        )
        service.run()
        repo = PublicationsRepo(db, conn)
        for pub_id in ("rep-sber", "rep-gazp", "ev-1"):
            row = repo.get_by_id(pub_id)
            assert row is not None and row.status == "written"


def test_writer_tickers_sheet_unions_db_with_config(
    tmp_path: Path,
) -> None:
    """Anti-regression: the ``tickers`` sheet must reflect every ticker
    the DB has ever recorded (so it stays consistent with the metrics
    sheet, which is DB-backed). Tickers known only to the DB — typical of
    synthetic ``EDX{id}`` tickers from evolve runs that re-write
    ``config-evolve/tickers.yaml`` to the current 3-ticker batch — get
    safe defaults: profile='non_bank', use_vision_extraction=False.
    Tickers known only to the active config (newly-added but not yet
    processed) ALSO appear so the prior contract is preserved."""
    db = Database(tmp_path / "state.sqlite")
    db.migrate()

    # DB has 4 issuers: 2 also in config (SBER/GAZP), 2 evolve-only
    # (EDX1021/EDX105 — synthetic tickers from prior evolve ticks).
    with closing(db.connect()) as conn:
        TickersRepo(db, conn).upsert_from_config(
            [
                TickerEntry(ticker="SBER", e_disclosure_id="1", name="Sberbank"),
                TickerEntry(ticker="GAZP", e_disclosure_id="2", name="Gazprom"),
                TickerEntry(
                    ticker="EDX1021",
                    e_disclosure_id="1021",
                    name="НОТА-Банк",
                ),
                TickerEntry(
                    ticker="EDX105",
                    e_disclosure_id="105",
                    name="НИКО-БАНК",
                ),
            ]
        )

    # Config has SBER + GAZP + LKOH (LKOH is configured but not yet in DB
    # — newly added).
    config = TickersConfig(
        tickers=[
            TickerEntry(
                ticker="SBER", e_disclosure_id="1", name="Sberbank",
                profile="bank",
            ),
            TickerEntry(
                ticker="GAZP", e_disclosure_id="2", name="Gazprom",
                profile="non_bank",
            ),
            TickerEntry(
                ticker="LKOH", e_disclosure_id="17", name="Lukoil",
                profile="non_bank", use_vision_extraction=True,
            ),
        ]
    )

    excel_path = tmp_path / "out.xlsx"
    with closing(db.connect()) as conn:
        WriterService(
            PublicationsRepo(db, conn),
            MetricsRepo(db, conn),
            EventsRepo(db, conn),
            QAIssuesRepo(db, conn),
            TickersRepo(db, conn),
            tickers_config=config,
            excel_path=excel_path,
        ).run()

    wb = load_workbook(excel_path)
    rows = list(wb["tickers"].iter_rows(values_only=True))
    header, *data = rows
    by_ticker = {r[0]: r for r in data}

    # All 4 DB tickers + 1 config-only ticker = 5 rows.
    assert set(by_ticker) == {"SBER", "GAZP", "EDX1021", "EDX105", "LKOH"}, (
        f"unexpected tickers on sheet: {set(by_ticker)}"
    )

    # Configured tickers carry their config-declared profile + vision flag.
    assert by_ticker["SBER"][2] == "bank"
    assert by_ticker["GAZP"][2] == "non_bank"
    assert by_ticker["LKOH"][2] == "non_bank"
    assert by_ticker["LKOH"][4] is True  # use_vision_extraction

    # Synthetic EDX-tickers known only to the DB get safe defaults.
    assert by_ticker["EDX1021"][2] == "non_bank"
    assert by_ticker["EDX1021"][4] is False
    assert by_ticker["EDX105"][2] == "non_bank"
    assert by_ticker["EDX105"][4] is False


def test_writer_idempotent_second_run(
    workspace: tuple[Database, Path],
) -> None:
    db, excel_path = workspace
    with closing(db.connect()) as conn:
        service = WriterService(
            PublicationsRepo(db, conn),
            MetricsRepo(db, conn),
            EventsRepo(db, conn),
            QAIssuesRepo(db, conn),
            TickersRepo(db, conn),
            tickers_config=_TICKERS_CONFIG,
            excel_path=excel_path,
        )
        first = service.run()
        # Second run should not raise even though publications are now 'written'
        # and there's nothing to flip.
        second = service.run()
    assert first == second == excel_path
    wb = load_workbook(excel_path)
    metrics = wb["metrics"]
    # Still 10 data rows after re-run.
    assert metrics.max_row == 11
