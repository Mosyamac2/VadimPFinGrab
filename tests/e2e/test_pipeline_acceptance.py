"""End-to-end acceptance tests (ТЗ §15 / prompt 15).

Runs real stage services against:
- ``httpx.MockTransport`` for the e-disclosure HTTP layer,
- a fake :class:`LLMProvider` returning canned JSON,
- a fake :class:`CloudStorageProvider` for Google Drive.

State DB and Excel writes are real.
"""

from __future__ import annotations

import json
import sqlite3
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import httpx
import pymupdf
import pytest
from openpyxl import load_workbook

from edx.config import (
    AppSettings,
    TickerEntry,
    load_all,
)
from edx.http.client import EDisclosureClient
from edx.orchestrator import Orchestrator, StageBundle
from edx.providers.llm import LLMRequest, LLMResponse
from edx.providers.storage import RemoteFileInfo
from edx.stages.classifier import build_classifier_service
from edx.stages.discoverer import build_discoverer_service
from edx.stages.downloader import build_downloader_service
from edx.stages.event_extractor import build_event_extractor_service
from edx.stages.metric_extractor import build_metric_extractor_service
from edx.stages.text_extractor import build_text_extractor_service
from edx.stages.unpacker import build_unpacker_service
from edx.stages.validator import build_validator_service
from edx.stages.writer import build_writer_service
from edx.stages.writer.replicator import ReplicatorService
from edx.storage import (
    Database,
    DocumentsRepo,
    EventsRepo,
    MetricsRepo,
    PublicationsRepo,
    QAIssuesRepo,
    RunsRepo,
    TickersRepo,
)

REPO_CONFIG_DIR = Path(__file__).resolve().parents[2] / "config"


# --------------------------------------------------------------------------
# Synthetic e-disclosure resources
# --------------------------------------------------------------------------


def _make_text_pdf(body: str) -> bytes:
    doc = pymupdf.open()
    page = doc.new_page(width=595, height=842)
    y = 50.0
    for line in body.splitlines() or [""]:
        page.insert_text(
            pymupdf.Point(50, y), line, fontsize=11, fontname="helv"
        )
        y += 14
    blob = doc.write()
    doc.close()
    return bytes(blob)


def _ifrs_report_body(ticker: str) -> str:
    # Latin only — Helvetica would silently drop Cyrillic. That's enough for
    # the Classifier to score IFRS markers + clear text content for the
    # Text Extractor.
    return (
        f"Issuer: {ticker}.\n"
        "IFRS Group consolidated financial statements 2025.\n"
        "Statement of Financial Position as at 31 December 2025.\n"
    ) * 12


_ISSUER_CARDS = {
    "1": (
        "<html><body>"
        "<section class='publications-section' data-section='reports'>"
        "<ul class='publications-list'>"
        "<li class='publication-row'>"
        "<span class='publication-date'>15.03.2026</span>"
        "<a class='publication-link' href='/portal/files/SBER/2025-fy.pdf'>"
        "MSFO 2025</a>"
        "</li></ul></section>"
        "<section class='publications-section' data-section='events'>"
        "<ul class='publications-list'>"
        "<li class='publication-row'>"
        "<span class='publication-date'>10.04.2026</span>"
        "<a class='publication-link' href='/portal/messages/SBER-divs.html'>"
        "Dividends decision</a>"
        "</li></ul></section>"
        "</body></html>"
    ),
    "2": (
        "<html><body>"
        "<section class='publications-section' data-section='reports'>"
        "<ul class='publications-list'>"
        "<li class='publication-row'>"
        "<span class='publication-date'>20.03.2026</span>"
        "<a class='publication-link' href='/portal/files/GAZP/2025-fy.pdf'>"
        "MSFO 2025</a>"
        "</li></ul></section>"
        "<section class='publications-section' data-section='events'>"
        "<ul class='publications-list'>"
        "<li class='publication-row'>"
        "<span class='publication-date'>12.04.2026</span>"
        "<a class='publication-link' href='/portal/messages/GAZP-mgmt.html'>"
        "Management change</a>"
        "</li></ul></section>"
        "</body></html>"
    ),
}

_EVENT_HTMLS = {
    "/portal/messages/SBER-divs.html": (
        "<html><body><main><article>"
        "<h1>Dividends decision</h1>"
        "<p>Sberbank announced dividends of 22.50 RUB per share for 2025.</p>"
        "</article></main></body></html>"
    ),
    "/portal/messages/GAZP-mgmt.html": (
        "<html><body><main><article>"
        "<h1>Management change</h1>"
        "<p>Gazprom appointed Ivan Ivanov as new CEO from April 1, 2026.</p>"
        "</article></main></body></html>"
    ),
}


# --------------------------------------------------------------------------
# Fakes
# --------------------------------------------------------------------------


@dataclass
class _FakeLLM:
    name: str = "fake_llm"
    supports_pdf_input: bool = False
    metric_calls: int = 0
    event_calls: int = 0

    @property
    def calls(self) -> int:
        return self.metric_calls + self.event_calls

    async def complete(self, req: LLMRequest) -> LLMResponse:
        props = req.json_schema.get("properties", {})
        if "extractions" in props:
            self.metric_calls += 1
            data: dict[str, Any] = {
                "extractions": [
                    {
                        "reporting_date": "2025-12-31",
                        "period_type": "FY",
                        "reporting_standard": "IFRS",
                        "currency": "RUB",
                        "unit": "ones",
                        "metrics": {
                            "revenue": {
                                "value": 1_000_000.0,
                                "source_quote": "Revenue 1000000",
                            },
                            "ebitda": {
                                "value": 300_000.0,
                                "source_quote": "EBITDA 300000",
                            },
                            "net_income": {
                                "value": 100_000.0,
                                "source_quote": "Net income 100000",
                            },
                            "total_assets": {
                                "value": 5_000_000.0,
                                "source_quote": "Total assets 5000000",
                            },
                            "total_debt": {
                                "value": 2_000_000.0,
                                "source_quote": "Total debt 2000000",
                            },
                        },
                    }
                ]
            }
        elif "event_type" in props:
            self.event_calls += 1
            data = {
                "event_type": "dividends",
                "event_date": "2026-04-10",
                "publication_date": "2026-04-10",
                "summary": "Dividends decision: 22.50 RUB per share.",
                "key_params": {"per_share_rub": 22.5},
            }
        else:
            raise AssertionError(f"unexpected schema: {sorted(props)}")

        return LLMResponse(
            data=data,
            raw_text=json.dumps(data, ensure_ascii=False),
            provider=self.name,
            model="fake-model",
            input_tokens=10,
            output_tokens=5,
        )


@dataclass
class _FakeCloudStorage:
    name: str = "fake_drive"
    calls: list[dict[str, Any]] = field(default_factory=list)

    def upsert_file(
        self,
        local_path: Path,
        remote_folder_id: str,
        remote_name: str,
        *,
        archive: bool,
    ) -> RemoteFileInfo:
        self.calls.append(
            {
                "local_path": Path(local_path),
                "remote_folder_id": remote_folder_id,
                "remote_name": remote_name,
                "archive": archive,
            }
        )
        return RemoteFileInfo(
            file_id="fake-file-id",
            web_view_link="https://drive.test/fake",
            updated_at="2026-05-01T00:00:00+00:00",
        )


# --------------------------------------------------------------------------
# Mock transport
# --------------------------------------------------------------------------


def _build_transport(
    *,
    pdf_bodies: dict[str, bytes],
    fail_paths: set[str] | None = None,
) -> httpx.MockTransport:
    fail_paths = fail_paths or set()

    def handler(request: httpx.Request) -> httpx.Response:
        path = request.url.path
        if path in fail_paths:
            return httpx.Response(500, text="boom")
        if path == "/robots.txt":
            return httpx.Response(200, text="")
        if path == "/portal/company.aspx":
            issuer_id = request.url.params.get("id", "")
            html = _ISSUER_CARDS.get(issuer_id, "<html><body></body></html>")
            return httpx.Response(
                200, text=html, headers={"Content-Type": "text/html"}
            )
        if path in pdf_bodies:
            return httpx.Response(
                200,
                content=pdf_bodies[path],
                headers={"Content-Type": "application/pdf"},
            )
        if path in _EVENT_HTMLS:
            return httpx.Response(
                200,
                content=_EVENT_HTMLS[path].encode("utf-8"),
                headers={"Content-Type": "text/html; charset=utf-8"},
            )
        return httpx.Response(404, text="not found")

    return httpx.MockTransport(handler)


# --------------------------------------------------------------------------
# Workspace
# --------------------------------------------------------------------------


@dataclass
class _Workspace:
    settings: AppSettings
    db: Database
    conn: sqlite3.Connection
    publications_repo: PublicationsRepo
    documents_repo: DocumentsRepo
    metrics_repo: MetricsRepo
    events_repo: EventsRepo
    qa_issues_repo: QAIssuesRepo
    runs_repo: RunsRepo
    tickers_repo: TickersRepo


def _make_workspace(tmp_path: Path) -> _Workspace:
    settings = load_all(REPO_CONFIG_DIR, env_file=tmp_path / "missing.env")
    settings.app.paths.data_dir = tmp_path / "data"
    settings.app.paths.raw_dir = tmp_path / "data" / "raw"
    settings.app.paths.processed_dir = tmp_path / "data" / "processed"
    settings.app.paths.state_db = tmp_path / "data" / "state.sqlite"
    settings.app.paths.output_dir = tmp_path / "output"
    settings.app.paths.excel_path = tmp_path / "output" / "e-disclosure.xlsx"
    settings.app.paths.logs_dir = tmp_path / "logs"
    settings.tickers.tickers = [
        TickerEntry(ticker="SBER", e_disclosure_id="1", name="Sberbank"),
        TickerEntry(ticker="GAZP", e_disclosure_id="2", name="Gazprom"),
    ]
    settings.app.discoverer.max_retries = 0
    settings.app.discoverer.retry_min_wait_s = 0.0
    settings.app.discoverer.retry_max_wait_s = 0.01
    settings.app.discoverer.requests_per_second = 100.0

    db = Database(settings.app.paths.state_db)
    db.migrate()
    conn = db.connect()
    publications_repo = PublicationsRepo(db, conn)
    documents_repo = DocumentsRepo(db, conn)
    metrics_repo = MetricsRepo(db, conn)
    events_repo = EventsRepo(db, conn)
    qa_issues_repo = QAIssuesRepo(db, conn)
    runs_repo = RunsRepo(db, conn)
    tickers_repo = TickersRepo(db, conn)
    tickers_repo.upsert_from_config(settings.tickers.tickers)

    return _Workspace(
        settings=settings,
        db=db,
        conn=conn,
        publications_repo=publications_repo,
        documents_repo=documents_repo,
        metrics_repo=metrics_repo,
        events_repo=events_repo,
        qa_issues_repo=qa_issues_repo,
        runs_repo=runs_repo,
        tickers_repo=tickers_repo,
    )


async def _run_pipeline(
    workspace: _Workspace,
    *,
    transport: httpx.MockTransport,
    fake_llm: _FakeLLM,
    fake_cloud: _FakeCloudStorage,
) -> Any:
    settings = workspace.settings
    async with EDisclosureClient(
        base_url=settings.app.discoverer.base_url,
        user_agent="edx-acceptance/1.0",
        requests_per_second=settings.app.discoverer.requests_per_second,
        request_timeout_s=settings.app.discoverer.request_timeout_s,
        max_retries=settings.app.discoverer.max_retries,
        retry_min_wait_s=settings.app.discoverer.retry_min_wait_s,
        retry_max_wait_s=settings.app.discoverer.retry_max_wait_s,
        respect_robots=False,
        transport=transport,
    ) as http_client:
        discoverer, _ = build_discoverer_service(
            settings, workspace.publications_repo, client=http_client
        )
        downloader = build_downloader_service(
            settings, workspace.publications_repo, client=http_client
        )
        unpacker = build_unpacker_service(
            settings,
            workspace.db,
            workspace.publications_repo,
            workspace.documents_repo,
        )
        classifier = build_classifier_service(
            settings, workspace.publications_repo, workspace.documents_repo
        )
        text_extractor = build_text_extractor_service(
            settings, workspace.publications_repo, workspace.documents_repo
        )
        metric_extractor = build_metric_extractor_service(
            settings,
            workspace.publications_repo,
            workspace.documents_repo,
            workspace.metrics_repo,
            fake_llm,  # type: ignore[arg-type]
        )
        event_extractor = build_event_extractor_service(
            settings,
            workspace.publications_repo,
            workspace.documents_repo,
            workspace.events_repo,
            fake_llm,  # type: ignore[arg-type]
        )
        validator = build_validator_service(
            settings,
            workspace.publications_repo,
            workspace.metrics_repo,
            workspace.qa_issues_repo,
        )
        writer = build_writer_service(
            settings,
            workspace.publications_repo,
            workspace.metrics_repo,
            workspace.events_repo,
            workspace.qa_issues_repo,
            workspace.tickers_repo,
        )
        replicator = ReplicatorService(
            provider=fake_cloud,  # type: ignore[arg-type]
            runs_repo=workspace.runs_repo,
            enabled=True,
            folder_id="acceptance-test-folder",
            file_name="e-disclosure.xlsx",
            archive=False,
        )
        bundle = StageBundle(
            discoverer=discoverer,
            downloader=downloader,
            unpacker=unpacker,
            classifier=classifier,
            text_extractor=text_extractor,
            metric_extractor=metric_extractor,
            event_extractor=event_extractor,
            validator=validator,
            writer=writer,
            replicator=replicator,
        )
        orchestrator = Orchestrator(
            runs_repo=workspace.runs_repo,
            publications_repo=workspace.publications_repo,
            metrics_repo=workspace.metrics_repo,
            events_repo=workspace.events_repo,
            qa_issues_repo=workspace.qa_issues_repo,
            stages=bundle,
            ticker_entries=settings.tickers.tickers,
            excel_path=settings.app.paths.excel_path,
            backfill_years=settings.app.mode.backfill_years,
        )
        return await orchestrator.run("update")


# --------------------------------------------------------------------------
# Scenario 1 — cold backfill
# --------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_scenario_1_cold_backfill(tmp_path: Path) -> None:
    workspace = _make_workspace(tmp_path)
    fake_llm = _FakeLLM()
    fake_cloud = _FakeCloudStorage()
    pdf_bodies = {
        "/portal/files/SBER/2025-fy.pdf": _make_text_pdf(
            _ifrs_report_body("SBER")
        ),
        "/portal/files/GAZP/2025-fy.pdf": _make_text_pdf(
            _ifrs_report_body("GAZP")
        ),
    }
    transport = _build_transport(pdf_bodies=pdf_bodies)

    try:
        outcome = await _run_pipeline(
            workspace,
            transport=transport,
            fake_llm=fake_llm,
            fake_cloud=fake_cloud,
        )
        publications = workspace.publications_repo.list_all()
        metrics = list(
            workspace.conn.execute("SELECT * FROM metrics ORDER BY metric_id")
        )
        events = list(
            workspace.conn.execute("SELECT * FROM events ORDER BY event_id")
        )
    finally:
        workspace.conn.close()

    assert outcome.status == "succeeded"
    statuses = {p.publication_id: p.status for p in publications}
    assert len(statuses) == 4
    assert all(status == "written" for status in statuses.values())

    assert len(metrics) == 10  # 5 metrics × 2 reports
    assert len(events) == 2

    # Excel mart written and matches the expected shape.
    excel_path = workspace.settings.app.paths.excel_path
    assert excel_path.is_file()
    wb = load_workbook(excel_path)
    assert {"metrics", "events", "meta", "qa_issues"}.issubset(set(wb.sheetnames))
    assert wb["metrics"].max_row == 11  # header + 10 rows
    assert wb["events"].max_row == 3  # header + 2 rows

    # Replicator received the local Excel path.
    assert len(fake_cloud.calls) == 1
    assert fake_cloud.calls[0]["local_path"] == excel_path
    assert fake_cloud.calls[0]["remote_name"] == "e-disclosure.xlsx"

    # LLM was hit exactly once per LLM-dependent publication.
    assert fake_llm.metric_calls == 2
    assert fake_llm.event_calls == 2


# --------------------------------------------------------------------------
# Scenario 2 — idempotent second run
# --------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_scenario_2_idempotent_second_run(tmp_path: Path) -> None:
    workspace = _make_workspace(tmp_path)
    fake_llm = _FakeLLM()
    fake_cloud = _FakeCloudStorage()
    pdf_bodies = {
        "/portal/files/SBER/2025-fy.pdf": _make_text_pdf(
            _ifrs_report_body("SBER")
        ),
        "/portal/files/GAZP/2025-fy.pdf": _make_text_pdf(
            _ifrs_report_body("GAZP")
        ),
    }
    transport = _build_transport(pdf_bodies=pdf_bodies)

    try:
        await _run_pipeline(
            workspace,
            transport=transport,
            fake_llm=fake_llm,
            fake_cloud=fake_cloud,
        )
        first_metric_calls = fake_llm.metric_calls
        first_event_calls = fake_llm.event_calls
        first_metric_count = workspace.conn.execute(
            "SELECT COUNT(*) AS c FROM metrics"
        ).fetchone()["c"]

        # Second run with the same DB state and same mock transport.
        outcome = await _run_pipeline(
            workspace,
            transport=transport,
            fake_llm=fake_llm,
            fake_cloud=fake_cloud,
        )

        runs = list(
            workspace.conn.execute("SELECT * FROM runs ORDER BY run_id")
        )
        metric_count = workspace.conn.execute(
            "SELECT COUNT(*) AS c FROM metrics"
        ).fetchone()["c"]
    finally:
        workspace.conn.close()

    assert outcome.status == "succeeded"
    # Second run finds 0 new publications.
    last_stats = json.loads(runs[-1]["stats_json"])
    assert last_stats["new_publications"] == 0
    # No additional LLM activity in the second pass.
    assert fake_llm.metric_calls == first_metric_calls
    assert fake_llm.event_calls == first_event_calls
    # Metrics row count is stable across runs.
    assert metric_count == first_metric_count


# --------------------------------------------------------------------------
# Scenario 3 — partial failure
# --------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_scenario_3_partial_failure_one_pdf_breaks(
    tmp_path: Path,
) -> None:
    workspace = _make_workspace(tmp_path)
    fake_llm = _FakeLLM()
    fake_cloud = _FakeCloudStorage()
    pdf_bodies = {
        "/portal/files/SBER/2025-fy.pdf": _make_text_pdf(
            _ifrs_report_body("SBER")
        ),
        "/portal/files/GAZP/2025-fy.pdf": _make_text_pdf(
            _ifrs_report_body("GAZP")
        ),
    }
    # The Downloader fails on the SBER report URL — the publication is
    # marked 'failed' and other 3 (GAZP report + 2 events) succeed.
    transport = _build_transport(
        pdf_bodies=pdf_bodies,
        fail_paths={"/portal/files/SBER/2025-fy.pdf"},
    )

    try:
        outcome = await _run_pipeline(
            workspace,
            transport=transport,
            fake_llm=fake_llm,
            fake_cloud=fake_cloud,
        )
        statuses = {
            p.publication_id: p.status
            for p in workspace.publications_repo.list_all()
        }
        qa_issues = workspace.qa_issues_repo.list_all()
    finally:
        workspace.conn.close()

    assert outcome.status == "partial"
    failed_ids = [pid for pid, status in statuses.items() if status == "failed"]
    written_ids = [pid for pid, status in statuses.items() if status == "written"]
    assert len(failed_ids) == 1
    assert len(written_ids) == 3

    # qa_issues contains a publication_failed entry for the failed publication.
    assert any(
        i.publication_id == failed_ids[0] and i.code == "publication_failed"
        for i in qa_issues
    )

    # Excel mart still exists with surviving publications.
    excel_path = workspace.settings.app.paths.excel_path
    assert excel_path.is_file()
    wb = load_workbook(excel_path)
    metrics_rows = wb["metrics"].max_row - 1  # exclude header
    # Only GAZP's 5 metrics — SBER's report never reached the LLM stage.
    assert metrics_rows == 5
