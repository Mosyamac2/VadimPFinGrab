"""Excel mart writer (ТЗ §10.3).

Atomic file creation:
- Workbook is saved to ``<target>.tmp`` and then ``os.replace``-d into place.
- Any failure mid-write removes the partial ``.tmp`` and leaves the prior
  file at ``<target>`` untouched (existing iPhone link stays valid — ТЗ §10.4).
"""

from __future__ import annotations

import contextlib
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

INTEGER_NUMBER_FORMAT: Final[str] = "#,##0"
DECIMAL_NUMBER_FORMAT: Final[str] = "#,##0.##"
MAX_AUTO_WIDTH: Final[int] = 60


@dataclass(frozen=True)
class MetricExportRow:
    ticker: str
    reporting_date: str
    period_type: str
    reporting_standard: str
    metric_name: str
    value: float | None
    currency: str
    unit: str
    qa_warning: str | None
    source_publication_url: str


@dataclass(frozen=True)
class EventExportRow:
    ticker: str
    event_date: str
    publication_date: str
    event_type: str
    summary: str
    key_params_json: str | None
    source_url: str


@dataclass(frozen=True)
class QAIssueExportRow:
    ticker: str
    publication_id: str
    code: str
    message: str
    created_at: str


@dataclass(frozen=True)
class MetaSnapshot:
    last_updated_at: str
    pipeline_version: str
    tickers_count: int
    metrics_rows: int
    events_rows: int
    incomplete_publications: int
    failed_publications: int


@dataclass(frozen=True)
class WitrineSnapshot:
    metrics: list[MetricExportRow] = field(default_factory=list)
    events: list[EventExportRow] = field(default_factory=list)
    qa_issues: list[QAIssueExportRow] = field(default_factory=list)
    meta: MetaSnapshot | None = None


METRICS_HEADERS: Final[tuple[str, ...]] = (
    "ticker",
    "reporting_date",
    "period_type",
    "reporting_standard",
    "metric_name",
    "value",
    "currency",
    "unit",
    "qa_warning",
    "source_publication_url",
)
EVENTS_HEADERS: Final[tuple[str, ...]] = (
    "ticker",
    "event_date",
    "publication_date",
    "event_type",
    "summary",
    "key_params_json",
    "source_url",
)
QA_ISSUES_HEADERS: Final[tuple[str, ...]] = (
    "ticker",
    "publication_id",
    "code",
    "message",
    "created_at",
)


class ExcelWriter:
    """Writes the four-sheet mart. Stateless — instantiated and used per-call."""

    def write(self, out_path: Path, snapshot: WitrineSnapshot) -> Path:
        wb = Workbook()
        # Replace the implicit empty sheet with our own four.
        default = wb.active
        if default is not None:
            wb.remove(default)

        self._write_metrics(wb, snapshot.metrics)
        self._write_events(wb, snapshot.events)
        self._write_meta(wb, snapshot.meta)
        self._write_qa_issues(wb, snapshot.qa_issues)

        out_path.parent.mkdir(parents=True, exist_ok=True)
        tmp = out_path.with_name(out_path.name + ".tmp")
        try:
            wb.save(tmp)
            os.replace(tmp, out_path)
        except BaseException:
            if tmp.exists():
                with contextlib.suppress(OSError):
                    tmp.unlink()
            raise
        return out_path

    def _write_metrics(
        self, wb: Workbook, rows: list[MetricExportRow]
    ) -> None:
        ws = wb.create_sheet("metrics")
        _write_headers(ws, METRICS_HEADERS)
        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row.ticker)
            ws.cell(row=row_idx, column=2, value=row.reporting_date)
            ws.cell(row=row_idx, column=3, value=row.period_type)
            ws.cell(row=row_idx, column=4, value=row.reporting_standard)
            ws.cell(row=row_idx, column=5, value=row.metric_name)
            value_cell = ws.cell(row=row_idx, column=6, value=row.value)
            if row.value is not None:
                if float(row.value).is_integer():
                    value_cell.number_format = INTEGER_NUMBER_FORMAT
                else:
                    value_cell.number_format = DECIMAL_NUMBER_FORMAT
            ws.cell(row=row_idx, column=7, value=row.currency)
            ws.cell(row=row_idx, column=8, value=row.unit)
            ws.cell(row=row_idx, column=9, value=row.qa_warning)
            ws.cell(row=row_idx, column=10, value=row.source_publication_url)
        _finalise_sheet(ws)

    def _write_events(
        self, wb: Workbook, rows: list[EventExportRow]
    ) -> None:
        ws = wb.create_sheet("events")
        _write_headers(ws, EVENTS_HEADERS)
        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row.ticker)
            ws.cell(row=row_idx, column=2, value=row.event_date)
            ws.cell(row=row_idx, column=3, value=row.publication_date)
            ws.cell(row=row_idx, column=4, value=row.event_type)
            ws.cell(row=row_idx, column=5, value=row.summary)
            ws.cell(row=row_idx, column=6, value=row.key_params_json)
            ws.cell(row=row_idx, column=7, value=row.source_url)
        _finalise_sheet(ws)

    def _write_meta(self, wb: Workbook, meta: MetaSnapshot | None) -> None:
        ws = wb.create_sheet("meta")
        ws.cell(row=1, column=1, value="key").font = Font(bold=True)
        ws.cell(row=1, column=2, value="value").font = Font(bold=True)
        if meta is not None:
            entries = [
                ("last_updated_at", meta.last_updated_at),
                ("pipeline_version", meta.pipeline_version),
                ("tickers_count", meta.tickers_count),
                ("metrics_rows", meta.metrics_rows),
                ("events_rows", meta.events_rows),
                ("incomplete_publications", meta.incomplete_publications),
                ("failed_publications", meta.failed_publications),
            ]
            for row_idx, (key, value) in enumerate(entries, start=2):
                ws.cell(row=row_idx, column=1, value=key)
                ws.cell(row=row_idx, column=2, value=value)
        _finalise_sheet(ws, freeze=False)

    def _write_qa_issues(
        self, wb: Workbook, rows: list[QAIssueExportRow]
    ) -> None:
        ws = wb.create_sheet("qa_issues")
        _write_headers(ws, QA_ISSUES_HEADERS)
        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row.ticker)
            ws.cell(row=row_idx, column=2, value=row.publication_id)
            ws.cell(row=row_idx, column=3, value=row.code)
            ws.cell(row=row_idx, column=4, value=row.message)
            ws.cell(row=row_idx, column=5, value=row.created_at)
        _finalise_sheet(ws)


def _write_headers(ws: Worksheet, headers: tuple[str, ...]) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)


def _finalise_sheet(ws: Worksheet, *, freeze: bool = True) -> None:
    if freeze:
        ws.freeze_panes = "A2"
    if ws.max_row == 0 or ws.max_column == 0:
        return
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 10
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            length = len(str(value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[letter].width = min(max_len + 2, MAX_AUTO_WIDTH)
